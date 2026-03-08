import openpyxl
from openpyxl.utils import get_column_letter
import akshare as ak
import time

# 股票名到代码映射（根据实际情况补充）
name_to_code = {
    "南山铝业": "600219",
    "江苏银行": "600919",
    # 可以继续补充
}

def get_latest_price_dividend(stock_name):
    """
    获取最新收盘价和近一年股息率
    返回 (最新价格, 最新股息率)
    """
    stock_code = name_to_code.get(stock_name)
    if not stock_code:
        print(f"{stock_name} 未找到对应股票代码，跳过")
        return None, None

    latest_price = None
    latest_div_yield = None

    # 获取最新价格
    try:
        stock_spot_df = ak.stock_zh_a_spot()
        price_row = stock_spot_df[stock_spot_df['代码'] == stock_code]
        if not price_row.empty:
            latest_price = float(price_row.iloc[0]['最新价'])
        else:
            print(f"{stock_name} 在 A 股行情列表中未找到，价格无法更新")
    except Exception as e:
        print(f"{stock_name} 获取最新价格失败: {e}")

    # 获取股息率
    try:
        dividend_df = ak.stock_dividend(stock_code)
        if not dividend_df.empty:
            possible_cols = ['股息率(%)', '分红收益率']
            col_name = next((c for c in possible_cols if c in dividend_df.columns), None)
            if col_name:
                latest_div_yield = float(dividend_df.iloc[-1][col_name]) / 100
            else:
                print(f"{stock_name} 分红表中未找到股息率列")
        else:
            print(f"{stock_name} 分红表为空，无法获取股息率")
    except Exception as e:
        print(f"{stock_name} 获取股息率失败: {e}")

    return latest_price, latest_div_yield


def main():
    file_path = r'C:\Users\hspcadmin\Desktop\股息目标测算表_年度20W.xlsx'

    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as e:
        print(f"加载工作簿出错：{e}")
        return

    ws = wb.active

    headers = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}

    required_cols = ['股数', '收盘价(元/股)', '成本元/股', '近一年股息率',
                     '当前市值/元', '持仓总成本(元)', '近一年预计分红(元)', '备注', '股名']
    for col in required_cols:
        if col not in headers:
            print(f"错误：缺少必要列 '{col}'")
            return

    data_rows = [r for r in range(2, ws.max_row + 1)
                 if ws.cell(row=r, column=headers['股数']).value is not None]

    if not data_rows:
        print("警告：未找到数据行，请检查文件内容。")
        return

    # 更新价格和股息率
    for row in data_rows:
        stock_name = str(ws.cell(row=row, column=headers['股名']).value).strip()
        latest_price, latest_div_yield = get_latest_price_dividend(stock_name)

        if latest_price is not None:
            ws.cell(row=row, column=headers['收盘价(元/股)'], value=latest_price)
        else:
            print(f"{stock_name} 收盘价未更新，保留原值")

        if latest_div_yield is not None:
            ws.cell(row=row, column=headers['近一年股息率'], value=latest_div_yield)
        else:
            print(f"{stock_name} 股息率未更新，保留原值")

        time.sleep(0.1)  # 减少请求压力

    # 写入公式
    col_shares = get_column_letter(headers['股数'])
    col_price = get_column_letter(headers['收盘价(元/股)'])
    col_cost = get_column_letter(headers['成本元/股'])
    col_div_yield = get_column_letter(headers['近一年股息率'])
    col_market = get_column_letter(headers['当前市值/元'])
    col_total_cost = get_column_letter(headers['持仓总成本(元)'])
    col_dividend = get_column_letter(headers['近一年预计分红(元)'])
    col_remark = get_column_letter(headers['备注'])

    for row in data_rows:
        ws.cell(row=row, column=headers['当前市值/元'], value=f"={col_shares}{row}*{col_price}{row}")
        ws.cell(row=row, column=headers['持仓总成本(元)'], value=f"={col_shares}{row}*{col_cost}{row}")
        ws.cell(row=row, column=headers['近一年预计分红(元)'], value=f"={col_shares}{row}*{col_price}{row}*{col_div_yield}{row}")

    # 汇总行公式
    summary_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value and '汇总' in str(ws.cell(r, 1).value):
            summary_row = r
            break

    if summary_row:
        first_data_row = min(data_rows)
        last_data_row = summary_row - 1
        ws.cell(row=summary_row, column=headers['当前市值/元'],
                value=f"=SUM({col_market}{first_data_row}:{col_market}{last_data_row})")
        ws.cell(row=summary_row, column=headers['持仓总成本(元)'],
                value=f"=SUM({col_total_cost}{first_data_row}:{col_total_cost}{last_data_row})")
        ws.cell(row=summary_row, column=headers['近一年预计分红(元)'],
                value=f"=SUM({col_dividend}{first_data_row}:{col_dividend}{last_data_row})")
        ws.cell(row=summary_row, column=headers['备注'],
                value=f'="总股息率: "&TEXT(IFERROR({col_dividend}{summary_row}/{col_total_cost}{summary_row},0),"0.00%")')

    try:
        wb.save(file_path)
        print(f"文件已成功更新（使用 AkShare 获取最新价格和股息率）：{file_path}")
    except Exception as e:
        print(f"保存文件时出错：{e}")


if __name__ == "__main__":
    main()