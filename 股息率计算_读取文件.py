import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import requests
import time

# 文件路径
mapping_file = r"C:\Users\hspcadmin\Desktop\A股股票代码名称.xlsx"
target_file = r"C:\Users\hspcadmin\Desktop\股息目标测算表_年度20W.xlsx"

# 读取中文名 -> 股票代码映射
mapping_df = pd.read_excel(mapping_file)
name_to_code = dict(zip(mapping_df['name'], mapping_df['code']))

def get_tencent_price(stock_code):
    """抓取腾讯接口最新价，自动识别字段"""
    url = f"http://qt.gtimg.cn/q={stock_code}"
    try:
        r = requests.get(url, timeout=5)
        r.encoding = 'gbk'
        text = r.text.strip()
        if not text or 'v_' not in text:
            print(f"{stock_code} 接口返回为空或格式错误: {text}")
            return None
        data = text.split('"')[1].split('~')
        # 尝试常见位置：3（现价）、2（昨收）、4（今开）
        for idx in [3, 2, 4]:
            try:
                price = float(data[idx])
                return price
            except:
                continue
        print(f"{stock_code} 未能解析最新价, 返回字段: {data}")
        return None
    except Exception as e:
        print(f"{stock_code} 获取最新价失败: {e}")
        return None

def main():
    wb = openpyxl.load_workbook(target_file, data_only=False)
    ws = wb.active

    # 获取列标题索引
    headers = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}

    required_cols = ['股名', '股数', '收盘价(元/股)', '成本元/股', '近一年分红/股(元)',
                     '近一年股息率', '当前市值/元', '持仓总成本(元)', '近一年预计分红(元)', '备注']
    for col in required_cols:
        if col not in headers:
            raise ValueError(f"缺少必要列: {col}")

    # 确定数据行（有股名的行，包括可能新插入的行）
    data_rows = [r for r in range(2, ws.max_row + 1)
                 if ws.cell(r, column=headers['股名']).value is not None]

    # 更新收盘价
    for row in data_rows:
        stock_name = ws.cell(row=row, column=headers['股名']).value.strip()
        stock_code = name_to_code.get(stock_name)
        if stock_code is not None:
            stock_code_str = str(stock_code).zfill(6)
            stock_code_prefixed = ('sh' if stock_code_str.startswith('6') else 'sz') + stock_code_str
            price = get_tencent_price(stock_code_prefixed)
            if price is not None:
                ws.cell(row=row, column=headers['收盘价(元/股)'], value=price)
            else:
                print(f"{stock_name} 收盘价获取失败，保留原值")
        else:
            print(f"{stock_name} 未找到对应股票代码")
        time.sleep(0.1)

    # 获取列字母（用于公式）
    col_shares = get_column_letter(headers['股数'])
    col_price = get_column_letter(headers['收盘价(元/股)'])
    col_cost = get_column_letter(headers['成本元/股'])
    col_div_yield = get_column_letter(headers['近一年股息率'])          # 手动输入的百分比列
    col_market = get_column_letter(headers['当前市值/元'])
    col_total_cost = get_column_letter(headers['持仓总成本(元)'])
    col_dividend = get_column_letter(headers['近一年预计分红(元)'])
    col_remark = get_column_letter(headers['备注'])

    # 为所有数据行写入公式（确保新行也获得公式）
    for row in data_rows:
        # 当前市值 = 股数 * 收盘价
        ws.cell(row=row, column=headers['当前市值/元'], value=f"={col_shares}{row}*{col_price}{row}")

        # 持仓总成本 = 股数 * 成本
        ws.cell(row=row, column=headers['持仓总成本(元)'], value=f"={col_shares}{row}*{col_cost}{row}")

        # 近一年预计分红 = 股数 * 收盘价 * 近一年股息率（百分比列）
        ws.cell(row=row, column=headers['近一年预计分红(元)'],
                value=f"={col_shares}{row}*{col_price}{row}*{col_div_yield}{row}")

    # 定位汇总行（A列包含“汇总”）
    summary_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value and '汇总' in str(ws.cell(r, 1).value):
            summary_row = r
            break

    if summary_row:
        # 数据区域从第2行到汇总行-1
        first_data_row = 2
        last_data_row = summary_row - 1
        # 汇总行公式（自动适应数据行范围）
        ws.cell(row=summary_row, column=headers['当前市值/元'],
                value=f"=SUM({col_market}{first_data_row}:{col_market}{last_data_row})")
        ws.cell(row=summary_row, column=headers['持仓总成本(元)'],
                value=f"=SUM({col_total_cost}{first_data_row}:{col_total_cost}{last_data_row})")
        ws.cell(row=summary_row, column=headers['近一年预计分红(元)'],
                value=f"=SUM({col_dividend}{first_data_row}:{col_dividend}{last_data_row})")
        ws.cell(row=summary_row, column=headers['备注'],
                value=f'="总股息率: "&TEXT(IFERROR({col_dividend}{summary_row}/{col_total_cost}{summary_row},0),"0.00%")')

    wb.save(target_file)
    print("Excel已更新：收盘价、市值、持仓成本、预计分红均已完成，股息率列保持手动输入")
    print("提示：新插入的行必须填写股名，运行脚本后会自动获得公式。若希望实时自动填充，请将数据区域转换为Excel表格（插入->表格）。")

if __name__ == "__main__":
    main()