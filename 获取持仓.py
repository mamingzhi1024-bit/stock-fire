"""
自动更新股息测算表脚本（同花顺交易窗口直连版）
功能：
1. 直接连接已打开的同花顺交易窗口，获取持仓（需要用户提供窗口标题/类名）
2. 若失败，则尝试通过 easytrader 常规方式连接
3. 最后尝试读取同花顺手动导出的CSV文件（备选）
4. 更新Excel表格：匹配股票代码更新现有行，插入新股，调整汇总公式
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter
import easytrader
import pandas as pd
import pywinauto
from datetime import datetime

# ==================== 配置区域 ====================
# 同花顺交易窗口的标题（请根据实际窗口修改，支持模糊匹配）
# 例如：'网上股票交易系统'、'国信证券'、'交易' 等
TRADE_WINDOW_TITLE = "网上股票交易系统"  # 请根据实际情况修改

# 交易窗口的类名（可选，如果标题能唯一确定可不填）
TRADE_WINDOW_CLASS = "#32770"  # 通常为对话框类名

# 同花顺客户端路径（如果直接连接窗口失败，会尝试启动客户端）
THS_CLIENT_PATH = r"E:\Soft\同花顺\同花顺\hexin.exe"  # 请修改为您的实际路径

# 同花顺导出CSV的默认路径（备选方案）
CSV_FALLBACK_PATH = r"C:\Users\hspcadmin\Desktop\持仓导出.csv"

# 目标Excel文件路径
EXCEL_PATH = r"C:\Users\hspcadmin\Desktop\股息目标测算表_年度20W.xlsx"

# 同花顺账号密码（如果客户端未记住密码，可能需要自动登录；建议提前手动登录）
THS_ACCOUNT = os.getenv("THS_ACCOUNT", "")      # 可不填
THS_PASSWORD = os.getenv("THS_PASSWORD", "")    # 可不填
# =================================================

def print_all_windows():
    """打印所有顶层窗口，帮助用户找到交易窗口的标题和类名"""
    print("当前所有顶层窗口：")
    windows = pywinauto.Desktop(backend="win32").windows()
    for w in windows:
        title = w.window_text()
        class_name = w.class_name()
        if title:  # 只打印有标题的窗口，减少输出
            print(f"标题: {title}, 类名: {class_name}")
    print("-" * 50)

def fetch_positions_from_ths():
    """
    尝试连接同花顺交易窗口并获取持仓
    """
    print("尝试通过 easytrader 连接同花顺交易窗口...")
    try:
        # 先打印窗口，让用户确认（首次运行时可取消注释）
        # print_all_windows()

        # 创建同花顺实例
        user = easytrader.use('ths')
        print("创建同花顺实例成功")

        # 尝试直接连接交易窗口（通过标题或类名）
        connected = False
        try:
            # 方式1：使用标题和类名连接
            if TRADE_WINDOW_CLASS and TRADE_WINDOW_TITLE:
                user.connect(class_name=TRADE_WINDOW_CLASS, title=TRADE_WINDOW_TITLE)
                print(f"使用类名 '{TRADE_WINDOW_CLASS}' 和标题 '{TRADE_WINDOW_TITLE}' 连接成功")
                connected = True
            elif TRADE_WINDOW_TITLE:
                user.connect(title=TRADE_WINDOW_TITLE)
                print(f"使用标题 '{TRADE_WINDOW_TITLE}' 连接成功")
                connected = True
        except Exception as e:
            print(f"直接连接交易窗口失败: {e}")

        # 方式2：如果未连接，尝试通过进程名连接
        if not connected:
            for proc in ['ths', 'xiadan', '同花顺']:
                try:
                    user.connect(proc)
                    print(f"通过进程 '{proc}' 连接成功")
                    connected = True
                    break
                except:
                    continue

        # 方式3：如果仍未连接，尝试启动客户端（需路径）
        if not connected and os.path.exists(THS_CLIENT_PATH):
            try:
                user.connect(THS_CLIENT_PATH)
                print(f"通过路径 '{THS_CLIENT_PATH}' 启动客户端成功")
                connected = True
            except Exception as e:
                print(f"通过路径启动失败: {e}")

        if not connected:
            raise Exception("所有连接方式均失败，请确保交易窗口已打开并正确配置标题")

        # 等待交易界面加载（可能需手动登录）
        import time
        time.sleep(3)

        # 如果客户端未登录，尝试自动登录（需账号密码）
        if THS_ACCOUNT and THS_PASSWORD:
            try:
                user.login(THS_ACCOUNT, THS_PASSWORD)
                print("自动登录完成")
            except Exception as e:
                print(f"自动登录失败，请手动登录: {e}")
                input("请手动登录交易窗口，然后按回车继续...")

        # 获取持仓
        positions = user.position
        print(f"原始持仓数据: {positions}")

        if not positions:
            print("同花顺返回空持仓，请检查是否已登录并持有股票")
            return None

        # 解析持仓数据（根据实际返回的字段名调整）
        result = []
        for item in positions:
            code = str(item.get('证券代码', item.get('股票代码', ''))).strip()
            if len(code) < 6:
                code = code.zfill(6)
            name = item.get('证券名称', item.get('股票名称', ''))
            shares = float(item.get('当前持仓', item.get('持股数量', 0)))
            cost_price = float(item.get('成本价', item.get('持仓成本', 0)))
            close_price = float(item.get('最新价', item.get('现价', 0)))

            result.append({
                'code': code,
                'name': name,
                'shares': shares,
                'cost_price': cost_price,
                'close_price': close_price
            })
        print(f"解析后得到 {len(result)} 条持仓记录")
        return result
    except Exception as e:
        print(f"同花顺客户端获取持仓失败: {e}")
        return None

def fetch_positions_from_csv(csv_path):
    """从同花顺导出的CSV文件中读取持仓（备选方案）"""
    print(f"尝试从CSV文件读取持仓: {csv_path}")
    try:
        df = pd.read_csv(csv_path, encoding='gbk')
        print(f"CSV列名: {df.columns.tolist()}")

        # 根据实际CSV列名调整映射
        code_col = '证券代码' if '证券代码' in df.columns else '股票代码'
        name_col = '证券名称' if '证券名称' in df.columns else '股票名称'
        shares_col = '持仓数量' if '持仓数量' in df.columns else '持股数量'
        cost_col = '成本价' if '成本价' in df.columns else '持仓成本'
        price_col = '最新价' if '最新价' in df.columns else '现价'

        result = []
        for _, row in df.iterrows():
            code = str(row[code_col]).strip()
            if len(code) < 6:
                code = code.zfill(6)
            name = row[name_col]
            shares = float(row[shares_col])
            cost_price = float(row[cost_col])
            close_price = float(row[price_col])

            result.append({
                'code': code,
                'name': name,
                'shares': shares,
                'cost_price': cost_price,
                'close_price': close_price
            })
        print(f"从CSV读取到 {len(result)} 条持仓记录")
        return result
    except Exception as e:
        print(f"读取CSV失败: {e}")
        return None

def fetch_positions():
    """综合获取持仓：优先同花顺交易窗口，失败则尝试 CSV"""
    positions = fetch_positions_from_ths()
    if positions is not None:
        return positions
    print("同花顺交易窗口获取失败，尝试使用 CSV 备选方案...")
    positions = fetch_positions_from_csv(CSV_FALLBACK_PATH)
    return positions

def update_excel(file_path, positions):
    """将持仓数据更新到Excel"""
    if not positions:
        print("没有持仓数据，Excel未更新")
        return

    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except FileNotFoundError:
        print(f"错误：Excel文件未找到 - {file_path}")
        return
    except Exception as e:
        print(f"加载Excel文件出错: {e}")
        return

    ws = wb.active

    # 获取标题行索引
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            headers[cell.value] = col

    required = ['股票代码', '股票名称', '股数', '成本元/股', '收盘价(元/股)', '近一年股息率',
                '当前市值/元', '持仓总成本(元)', '近一年预计分红(元)', '备注']
    for name in required:
        if name not in headers:
            print(f"错误：Excel缺少必要列 '{name}'")
            return

    # 定位汇总行（A列包含“汇总”）
    summary_row = None
    for row in range(2, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1).value
        if cell_a and '汇总' in str(cell_a):
            summary_row = row
            break
    if summary_row is None:
        print("错误：未找到A列包含“汇总”的行")
        return

    # 建立现有股票代码到行号的映射
    existing_codes = {}
    for row in range(2, summary_row):
        code_cell = ws.cell(row=row, column=headers['股票代码']).value
        if code_cell:
            existing_codes[str(code_cell).strip()] = row

    # 分离更新和插入
    to_update = []   # (row, position)
    to_insert = []   # position
    for pos in positions:
        code = pos['code']
        if code in existing_codes:
            to_update.append((existing_codes[code], pos))
        else:
            to_insert.append(pos)

    print(f"待更新: {len(to_update)} 行，待新增: {len(to_insert)} 行")

    # 1. 更新现有行
    for row, pos in to_update:
        ws.cell(row=row, column=headers['股数'], value=pos['shares'])
        ws.cell(row=row, column=headers['成本元/股'], value=pos['cost_price'])
        ws.cell(row=row, column=headers['收盘价(元/股)'], value=pos['close_price'])
        # 股息率列不覆盖

    # 2. 插入新行
    if to_insert:
        ws.insert_rows(idx=summary_row, amount=len(to_insert))
        new_summary_row = summary_row + len(to_insert)
        for i, pos in enumerate(to_insert):
            new_row = summary_row + i
            ws.cell(row=new_row, column=headers['股票代码'], value=pos['code'])
            ws.cell(row=new_row, column=headers['股票名称'], value=pos['name'])
            ws.cell(row=new_row, column=headers['股数'], value=pos['shares'])
            ws.cell(row=new_row, column=headers['成本元/股'], value=pos['cost_price'])
            ws.cell(row=new_row, column=headers['收盘价(元/股)'], value=pos['close_price'])
            ws.cell(row=new_row, column=headers['近一年股息率'], value=0)  # 默认0

            # 写入公式
            col_shares = get_column_letter(headers['股数'])
            col_price = get_column_letter(headers['收盘价(元/股)'])
            col_cost = get_column_letter(headers['成本元/股'])
            col_div = get_column_letter(headers['近一年股息率'])

            ws.cell(row=new_row, column=headers['当前市值/元'], value=f"={col_shares}{new_row}*{col_price}{new_row}")
            ws.cell(row=new_row, column=headers['持仓总成本(元)'], value=f"={col_shares}{new_row}*{col_cost}{new_row}")
            ws.cell(row=new_row, column=headers['近一年预计分红(元)'], value=f"={col_shares}{new_row}*{col_price}{new_row}*{col_div}{new_row}")

        summary_row = new_summary_row

    # 3. 更新汇总行公式
    first_data_row = 2
    last_data_row = summary_row - 1
    col_market = get_column_letter(headers['当前市值/元'])
    col_total_cost = get_column_letter(headers['持仓总成本(元)'])
    col_dividend = get_column_letter(headers['近一年预计分红(元)'])

    ws.cell(row=summary_row, column=headers['当前市值/元'], value=f"=SUM({col_market}{first_data_row}:{col_market}{last_data_row})")
    ws.cell(row=summary_row, column=headers['持仓总成本(元)'], value=f"=SUM({col_total_cost}{first_data_row}:{col_total_cost}{last_data_row})")
    ws.cell(row=summary_row, column=headers['近一年预计分红(元)'], value=f"=SUM({col_dividend}{first_data_row}:{col_dividend}{last_data_row})")

    # 20w目前完成度（如果存在）
    if '20w目前完成度' in headers:
        col_completion = get_column_letter(headers['20w目前完成度'])
        ws.cell(row=summary_row, column=headers['20w目前完成度'], value=f"={col_dividend}{summary_row}/200000")

    # 备注列的总股息率公式
    remark_formula = f'="总股息率: "&TEXT(IFERROR({col_dividend}{summary_row}/{col_total_cost}{summary_row},0),"0.00%")'
    ws.cell(row=summary_row, column=headers['备注'], value=remark_formula)

    # 保存文件
    try:
        wb.save(file_path)
        print(f"Excel文件已成功更新：{file_path}")
    except Exception as e:
        print(f"保存Excel文件失败：{e}")

def main():
    print(f"开始执行更新任务：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    # 首次运行可取消下面注释，打印所有窗口，帮助填写配置
    # print_all_windows()
    positions = fetch_positions()
    if positions:
        update_excel(EXCEL_PATH, positions)
    else:
        print("所有获取持仓的方式均失败，表格未更新。")

if __name__ == "__main__":
    main()